from django.shortcuts import render, redirect
from django.contrib.auth.decorators import user_passes_test, login_required
from . models import Assignment, Assignment_Problems, Course, Course_Enrollement, Platform, Problem, ProblemTags, User_Handle, Solved_Problem, Profile, Batch, Institution, Department
from django.views.decorators.csrf import csrf_exempt
import requests
import json
import leetcode as lc
from django.http import JsonResponse
from openpyxl import load_workbook
from django.contrib.auth.models import User
from datetime import date
from django.template.loader import render_to_string
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm


# Create your views here.
def check_admin(user):
   return user.is_superuser

def check_coordinator(user):
    return user.groups.filter(name='coordinator').exists()


@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Your password was successfully updated!')
            return redirect('home')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'decodeschool/change_password.html', {
        'form': form
    })

@user_passes_test(check_admin)
def get_all_problems(request):
    
    url = "https://leetcode.com/graphql"
    data = {"query":"query problemsetQuestionList($categorySlug: String, $limit: Int, $skip: Int, $filters: QuestionListFilterInput) {\n  problemsetQuestionList: questionList(\n    categorySlug: $categorySlug\n    limit: $limit\n    skip: $skip\n    filters: $filters\n  ) {\n    total: totalNum\n    questions: data {\n      acRate\n      difficulty\n      freqBar\n      frontendQuestionId: questionFrontendId\n      isFavor\n      paidOnly: isPaidOnly\n      status\n      title\n      titleSlug\n      topicTags {\n        name\n        id\n        slug\n      }\n      hasSolution\n      hasVideoSolution\n    }\n  }\n}","variables":{"categorySlug":"","skip":0,"limit":2968,"filters":{}}}
    headers = {'Content-type': 'application/json', 'Accept': 'text/plain'}
    r = requests.get(url, data=json.dumps(data), headers=headers)
    response = r.json()
    platform = Platform.objects.get(name='Leet Code')
    for question in response['data']['problemsetQuestionList']['questions']:
        if not Problem.objects.filter(platform=platform,problem_slug=question['titleSlug']).exists():
            problem = Problem(problem_title=question['title'], platform=platform, problem_slug=question['titleSlug'], complexity= question['difficulty'])
            problem.save()

            for topic in question['topicTags']:
                problem_tag = ProblemTags(problem=problem,topictags=topic['name'])
                problem_tag.save()
    return redirect('home')

@login_required
def view_dashboard(request):
  
    user_handles = User_Handle.objects.filter(user=request.user)
    profile = Profile.objects.get(user=request.user)
    return render(request, 'decodeschool/dashboard.html', {'user_handles':user_handles, 'profile':profile})

@user_passes_test(check_coordinator)
def myadmin(request):
    profiles = sorted(Profile.objects.all(),  key=lambda a: a.get_count,reverse=True)
    institutions = Institution.objects.all().order_by('name')
    courses_a = Course.objects.all().order_by('course_title')
    assigments = Assignment.objects.all().order_by('-id')
    return render(request, 'decodeschool/admin.html', {'profiles':profiles,'institutions':institutions, 'courses_a':courses_a, 'assignments':assigments })

@user_passes_test(check_coordinator)
def get_assignment(request):
    data = dict()
    assignment_id = request.GET.get('assignment_id')
    institute_id = request.GET.get('institute_id')
    course_id = request.GET.get('course_id')

    

    if institute_id == "*" and course_id == "*":
        profiles = Profile.objects.all()
    elif institute_id == "*":
        profiles = Profile.objects.filter(user__course_enrollement__course__id=course_id)
    elif course_id == "*":
        profiles = Profile.objects.filter(institute__id=institute_id)
    else:
        profiles = Profile.objects.filter(user__course_enrollement__course__id=course_id,institute__id=institute_id)

    assignment_problems = Assignment_Problems.objects.filter(assignment__id=assignment_id)
    data['html'] = render_to_string('decodeschool/assignment.html',{'profiles':profiles,'assignment_problems':assignment_problems,'assignment_id':assignment_id},request)
    return JsonResponse(data)


@login_required
def get_lc_problems(request,pk):
    solved_problems = Solved_Problem.objects.filter(user__username=pk).order_by("-id")
    print(solved_problems)
    return render(request, 'decodeschool/problems.html', {'solved_problems':solved_problems})

@user_passes_test(check_coordinator)
def filter_institute(request):
    pk = request.GET.get('institute')
    if pk == "*":
        profiles = sorted(Profile.objects.all(),  key=lambda a: a.get_count,reverse=True)
    else:
        profiles = sorted(Profile.objects.filter(institute__id=pk),  key=lambda a: a.get_count,reverse=True)
        
    data = dict()
    data['html'] = render_to_string('decodeschool/profile_list.html', {'profiles':profiles}, 
                           request)
    return JsonResponse(data)

@login_required
@csrf_exempt
def update_leet(request,uh):
    data = dict()
    data["status"] = "not ok"
    try:

        token = request.POST.get('lc_csrf')
        leetcode_session = request.POST.get('lc_session')
        configuration = lc.Configuration()
        configuration.api_key["x-csrftoken"] = token
        configuration.api_key["csrftoken"] = token
        configuration.api_key["LEETCODE_SESSION"] = leetcode_session
        configuration.api_key["Referer"] = "https://leetcode.com"
        configuration.debug = False

        api_instance =lc.DefaultApi(lc.ApiClient(configuration))

        graphql_request = lc.GraphqlQuery(
        query="""
        {
            user {
            username
            isCurrentUserPremium
            }
        }
        """,
        variables=lc.GraphqlQueryVariables(),
        )
        response = api_instance.graphql_post(body=graphql_request)

        if response.data.user.username == uh:
            api_response=api_instance.api_problems_topic_get(topic="algorithms")
            #solved_questions=[]
            for questions in api_response.stat_status_pairs:

                if questions.status=="ac":
                    #solved_questions.append(questions.stat.question__title)
                    try:

                        problem = Problem.objects.get(problem_title=questions.stat.question__title)
                        if not Solved_Problem.objects.filter(user=request.user,problem=problem).exists():
                            solved_problem = Solved_Problem(user=request.user,problem=problem)
                            solved_problem.save()
                    except:
                        pass
                data['message'] = "Updated your submissions"
                data["status"] = "ok"
                handle = User_Handle.objects.get(platform__name="Leet Code", user=request.user)
                print(handle)
                handle.last_update = date.today()
                handle.save()
        else:
            data['message'] = "Wrong CSRF or Session ID, Check Cookies.."
    except Exception as e:
        data['message'] = e

        
        
    return JsonResponse(data)

@user_passes_test(check_admin)
def import_from_excel(request):
   
    data = dict()
    if request.method == 'POST':
        
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        
        ws = wb.active
        
        data['count'] = 0
        data['missed'] = ""
        
        if(ws.max_column==2):
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:

                    username,course = row
                    if not Course_Enrollement.objects.filter(user=User.objects.get(username=username),course=Course.objects.get(course_title=course)).exists():

                        Course_Enrollement.objects.create(user=User.objects.get(username=username),course=Course.objects.get(course_title=course))
                    data['count'] += 1
                except:
                    data['missed'] += username + " "
        else:
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                
                username, first_name,last_name,email,mobile,department,batch,institution,leetcode = row
                try:
                    user = User.objects.create(username=username, first_name=first_name,last_name=last_name,password=make_password("Karpagam@105"),email=email)
                    Profile.objects.create(mobile_no=mobile,email=email,user=user,batch=Batch.objects.get(end_year=batch),department=Department.objects.get(short_name=department),institute=Institution.objects.get(name=institution))
                    User_Handle.objects.create(platform=Platform.objects.get(name="Leet Code"),user_handle=leetcode,user=user)
                    data['count'] += 1
                
                except Exception as e:
                    
                    data['missed'] += username + " "
              
    return JsonResponse(data)